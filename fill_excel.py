# -*- coding: utf-8 -*-
import os
import openpyxl
import random
import sys

sys.stdout.reconfigure(encoding='utf-8')

# ============ 虚假数据池 ============
names = ['王明', '李华', '赵雪', '孙伟', '周杰', '吴芳', '郑强', '陈静', '林涛', '刘洋',
         '张萌', '杨帆', '黄磊', '许晴', '何晨', '高阳', '罗浩', '郭敏', '梁栋', '谢辉',
         '马超', '韩静', '胡磊', '朱琳', '诸葛浩', '司马晨', '欧阳文', '上官婷', '令狐飞', '公孙龙']

colleges = ['计算机学院', '电子信息工程学院', '自动化科学与电气工程学院', '机械工程及自动化学院',
            '材料科学与工程学院', '经济管理学院', '人文社会科学学院', '宇航学院', '飞行学院', '软件学院']

art_teams = ['北航学生合唱团', '北航学生交响乐团', '北航学生舞蹈团', '北航学生话剧团', 
             '北航学生民乐团', '北航学生京剧团', '北航学生朗诵团']

positions = ['团干', '团长', '副团长', '声部长', '首席', '演员', '乐手']

event_types = ['毕业生晚会', '迎新晚会', '校庆晚会', '新年音乐会', '校园文化艺术节', '其他']

locations = ['校内', '校外']

work_types_actor = ['演员', '合唱人员', '舞蹈演员', '演奏人员']
work_types_staff = ['工作人员', '道具组', '灯光组', '音响组', '舞美组']

competitions = [
    ('全国大学生艺术展演', 2015, '国家级'),
    ('北京市大学生音乐节', 2016, '省部级'),
    ('北京大学生舞蹈节', 2017, '省部级'),
    ('北京大学生戏剧节', 2018, '省部级'),
    ('中华号角-上海之春国际音乐节管乐艺术节', 2019, '国际级'),
    ('首都高校艺术展演', 2020, '省部级'),
    ('全国大学生合唱比赛', 2021, '国家级'),
    ('北京市高校戏剧比赛', 2022, '省部级'),
    ('全国艺术展演', 2023, '国家级'),
    ('北京市朗诵艺术节', 2023, '省部级'),
    ('全国大学生民乐比赛', 2022, '国家级'),
    ('北京市交响乐展演', 2021, '省部级'),
]

# ============ 生成虚假数据 ============
def generate_fake_student(idx):
    """生成虚假学生信息"""
    return {
        'name': random.choice(names),
        'student_id': f'21{random.randint(370000, 379999)}',
        'college': random.choice(colleges),
        'category': random.choice(['本科生', '硕士研究生', '博士研究生']),
        'art_team': random.choice(art_teams),
        'year_in': random.choice(['2020-2021学年', '2021-2022学年', '2022-2023学年']),
        'year_out': random.choice(['2021-2022学年', '2022-2023学年', '2023-2024学年']),
        'position': random.choice(positions),
        'year_position': random.choice(['2020-2021学年', '2021-2022学年', '2022-2023学年']),
    }

def generate_art_team_data(n):
    """生成艺术团履历数据"""
    data = []
    for i in range(n):
        stu = generate_fake_student(i)
        leader_name = random.choice(names)
        data.append([
            '',  # 序号
            stu['name'],
            stu['student_id'],
            stu['college'],
            stu['category'],
            stu['art_team'],
            stu['year_in'],
            stu['position'],
            stu['year_position'],
            leader_name,
            '图片'
        ])
    return data

def generate_competition_data(n):
    """生成竞赛奖项数据"""
    data = []
    for i in range(n):
        stu = generate_fake_student(i)
        comp = random.choice(competitions)
        data.append([
            '',  # 序号
            stu['name'],
            stu['student_id'],
            stu['college'],
            stu['category'],
            comp[1],  # 获奖年份
            comp[0],  # 竞赛名称
            f'作品{random.randint(1, 99)}',
            random.choice(['国家级', '省部级', '校级']),
            random.choice(['一等奖', '二等奖', '三等奖', '优秀奖']),
            random.choice(['团体', '个人']),
            '图片'
        ])
    return data

def generate_activity_data(n):
    """生成活动演出数据"""
    data = []
    for i in range(n):
        stu = generate_fake_student(i)
        is_actor = random.choice([True, False])
        data.append([
            '',  # 序号
            stu['name'],
            stu['student_id'],
            stu['college'],
            stu['category'],
            random.randint(2022, 2024),  # 活动年份
            random.choice(event_types),
            f'"{random.choice(["闪闪发光", "青春绽放", "梦想启航", "星耀北航", "启航新时代", "青春无悔"])}"2024年度文艺汇演',
            random.choice(locations),
            '演员' if is_actor else '工作人员',
            random.choice(work_types_actor if is_actor else work_types_staff),
            '图片'
        ])
    return data

# ============ 写入Excel ============
desktop = 'C:/Users/a4253/Desktop/'
target_file = os.path.join(desktop, '文艺类（三个sheet）-第二课堂成绩单数据收集.xlsx')

print(f"正在处理文件: {target_file}")
print("="*60)

wb = openpyxl.load_workbook(target_file)

# Sheet 1: 艺术团履历
ws1 = wb['艺术团履历']
art_data = generate_art_team_data(15)
start_row = 7  # 从第7行开始写入
for i, row_data in enumerate(art_data):
    for j, value in enumerate(row_data):
        ws1.cell(row=start_row + i, column=j + 1, value=value)
print(f"✓ 艺术团履历: 写入 {len(art_data)} 条记录")

# Sheet 2: 竞赛奖项
ws2 = wb['竞赛奖项']
comp_data = generate_competition_data(15)
start_row = 8  # 从第8行开始写入
for i, row_data in enumerate(comp_data):
    for j, value in enumerate(row_data):
        ws2.cell(row=start_row + i, column=j + 1, value=value)
print(f"✓ 竞赛奖项: 写入 {len(comp_data)} 条记录")

# Sheet 3: 活动演出
ws3 = wb['活动演出']
act_data = generate_activity_data(15)
start_row = 10  # 从第10行开始写入
for i, row_data in enumerate(act_data):
    for j, value in enumerate(row_data):
        ws3.cell(row=start_row + i, column=j + 1, value=value)
print(f"✓ 活动演出: 写入 {len(act_data)} 条记录")

# 保存文件
wb.save(target_file)
print("="*60)
print(f"✓ 文件已保存: {target_file}")
print("\n生成的虚假数据预览:")
print("\n【艺术团履历】")
for row in art_data:
    print(f"  {row[1]} | {row[2]} | {row[3]} | {row[5]} | {row[7]}")

print("\n【竞赛奖项】")
for row in comp_data:
    print(f"  {row[1]} | {row[2]} | {row[6]} | {row[9]}")

print("\n【活动演出】")
for row in act_data:
    print(f"  {row[1]} | {row[2]} | {row[6]} | {row[8]}")
